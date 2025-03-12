#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time    : 2024-01-02 15:45
# @Author  : Yanjun Wang
# @Site    : ${晴飔 浮阳}
# @File    : 回路法.py
# @Software: PyCharm
import openpyxl
import numpy as np

class Mine_Network:
    # 定义整数类型的变量
    def __init__(self):
        self.Branch_Count = 0       # *边数量
        self.Junction_Count = 0     # *顶点数量 
        self.Circuit_Count = 0      # *回路数量
        self.Fan_Count = 0          # *风机数量
        self.RequiredQ_Count = 0    # *所需风量

        # 定义单精度浮点数数组
        self.A = [0.0, 0.0, 0.0]  # 风机特性曲线的三个系数  # *从左到右依次为二次项系数、一次项系数、常数系数

        # 定义整数和字符串数组
        self.VertexStart = []       # *起点
        self.VertexEnd = []         # *终点
        self.Fan_Modals = []        # *风机？

        # 定义单精度浮点数数组
        self.AirQuantities = []     # *风量
        self.Resistances = []       # *阻力

        # 定义整数数组
        self.cn = []                
        self.BranchCountInCircuit = []      # *回路中边的数量
        self.BranchNumbersInCircuit = []    # *回路中边的编号

        # 定义单精度浮点数变量=
        self.Q0 = 20  # 风机的最佳风量


    def read(self):
        # 计算分支和回路数量
        self.shtBranches = openpyxl.load_workbook('/Users/xavier/Documents/Py/找回路等6个文件/实例网络解算.xlsm')   # *读取Excel表格
        self.Branch_Count = len(self.shtBranches['Branch']['A']) - 1# 分支条数，列A的非空单元格数减1 # *边的数量（生成树树枝数）
        self.Circuit_Count = len(self.shtBranches['Circuit']['A'])  # 回路个数，列A的非空单元格数   # *网孔数量（网孔数 = 边数 - 顶点数 + 1）   
        self.Junction_Count = self.Branch_Count - self.Circuit_Count + 1  # 节点个数，分支条数减去回路个数加1
        print(f"边的数量：{self.Branch_Count}\n网孔数量：{self.Circuit_Count}\n顶点个数：{self.Junction_Count}")    
        self.Fan_Count = 1  # 风机台数，固定为1
        self.RequiredQ_Count = 0  # 需风点数，固定为0
        # VertexStart, VertexEnd = shtBranches['Branch']['B']
        for i in self.shtBranches['Branch']['B']:       # *遍历Excel表格中的Branch表的B列
            self.VertexStart.append(i.internal_value)   # *读取起点编号
        self.VertexStart = self.VertexStart[1:]         # *去掉表中的第一行（删掉“起点”）
        for i in self.shtBranches['Branch']['C']:       # *同上
            self.VertexEnd.append(i.internal_value)
        self.VertexEnd = self.VertexEnd[1:]
            # , shtBranches['Branch']['C'].internal_value  # 分支起止点编号，列B和列C的值
        print(f"起点编号：{self.VertexStart}\n终点编号：{self.VertexEnd}")
        for i in self.shtBranches['Branch']['H']:   # 分支风阻，列H的值
            self.Resistances.append(i.internal_value)
        self.Resistances=self.Resistances[1:]
        self.AirQuantities = [0] * self.Branch_Count  # 分支风量，初始化为0的列表
        print(f"风量初始化：{self.AirQuantities}\n")
        # 计算回路中的分支编号和分支条数
        for i in range(1, self.Circuit_Count + 1):  # *遍历通风网络中的回路
            self.BranchNumbersCircuit = []  # *初始化一个空列表，用于储存当前回路的边（分支）的信息
            self.BranchCountInCircuit.append(0)  # 初始化回路中的分支条数计数为0
            for j in range(1, self.Junction_Count + 1): # *遍历该回路的每个顶点
                if self.shtBranches['Circuit'].cell(row=i, column=j).value == None: # *当读取到空单元格时，表示该回路中的顶点已被全部读取，随后结束循环
                    break
                self.BranchNumbersCircuit.append(self.shtBranches['Circuit'].cell(row=i, column=j).value)   # *读取当前回路中边的编号
            print(f"回路{i}:", self.BranchNumbersCircuit, "\n")
            # *将每个回路的边的信息列表储存到一个大的列表中（用一个列表储存多个列表）
            self.BranchNumbersInCircuit.append(self.BranchNumbersCircuit) # 设置回路中的分支编号信息    
            self.BranchCountInCircuit[i - 1] = j - 1  # 计算并设置回路中的分支条数
        print(f"全部回路:", self.BranchNumbersInCircuit, "\n")

        # 设置风机特性曲线的系数信息（这里没有给出具体实现）
        # A[0] = float(input("请输入风机特性曲线的第一个系数："))  # 用户输入第一个系数值并转换为浮点数类型
        # A[1] = float(input("请输入风机特性曲线的第二个系数："))  # 用户输入第二个系数值并转换为浮点数类型
        # A[2] = float(input("请输入风机特性曲线的第三个系数："))  # 用户输入第三个系数值并转换为浮点数类型
        self.A[0] = -0.55
        self.A[1] = 0.55
        self.A[2] = 1234

        print("读取原始数据完成！")


    # 初始化风量
    def initQ(self):
        i, j, k = 0, 0, 0   # *i表示回路索引、j表示回路中的边索引、k表示边的编号
        # shtBranches = openpyxl.load_workbook('实例网络解算.xlsm')
        if self.shtBranches['Branch'].cell(2, 9).value == 0:  # 假设 shtBranches 是一个 openpyxl 工作表对象
            for i in range(self.Circuit_Count):     # *遍历回路
                self.AirQuantities[self.BranchNumbersInCircuit[i][0]-1] = self.Q0
                for j in range(1,self.BranchCountInCircuit[i]): # *遍历当前回路的每条边
                    k = abs(self.BranchNumbersInCircuit[i][j])-1    
                    self.AirQuantities[k] = self.AirQuantities[k] + np.sign(self.BranchNumbersInCircuit[i][j]) * self.Q0    # *为每个回路中的各条边分配初始风量
            for i in range(self.Branch_Count):
                self.shtBranches['Branch'].cell(i + 1, 9).value = self.AirQuantities[i]  # 假设 AirQuantities 是一个字典
        else:
            for i in range(self.Branch_Count):
                self.AirQuantities[i] = self.shtBranches['Branch'].cell(i + 2, 9).value  # 假设 AirQuantities 是一个字典


    # 计算回路风量不平衡值
    def circuitF(self, cn):     # *cn为回路编号
        j, s, cF = 0, 0, 0.0    # *j是计数器，用于遍历回路中的每条分支  s表示边风量的方向（正负号） cF风压不平衡值
        for j in range(self.BranchCountInCircuit[cn]):  # *遍历回路中的每条边
            k = abs(self.BranchNumbersInCircuit[cn][j])
            s = np.sign(self.BranchNumbersInCircuit[cn][j])
            cF = cF+s * self.Resistances[k-1] * abs(self.AirQuantities[k-1]) * self.AirQuantities[k-1]
        return cF


    # 计算回路风量不平衡值的导数
    def circuitF1(self, cn):
        j,ff = 0,0.0
        for j in range(self.BranchCountInCircuit[cn]):
            k = abs(self.BranchNumbersInCircuit[cn][j])
            ff += 2 * self.Resistances[k-1] * abs(self.AirQuantities[k-1])
        return ff


    # 计算风机风压
    def fan(self,Q):
        return self.A[0] * Q ** 2 + self.A[1] * Q + self.A[2]


    # 计算风机风压导数
    def fan1(self,Q):
        return 2 * self.A[0] * Q + self.A[1]

    def make_circuits(self):    # *构建生成树
        self.Blocks = list(range(self.Junction_Count))  # 假设Junction_Count是一个已定义的整数 #* 用于跟踪节点属于哪个连通块，初识时，每个节点属于自己的块
        self.Tree = [0]*(self.Junction_Count-1) # *初始化生成树中边的数量 （self.Junction_Count-1即为顶点数减1）
        n = 0   # *计数生成树的边数
        # *遍历图中的所有边
        for i in range(self.Branch_Count):  # 假设Branch_Count是一个已定义的整数 
            if self.Blocks[self.VertexStart[i]-1] != self.Blocks[self.VertexEnd[i]-1]:  # 假设VertexStart和VertexEnd是返回索引的函数    #* 如果起点和终点不形成回路
                self.Tree.append(i) # *加入生成树的边列表
                n += 1
            else:
                self.find_circuit(i, self.Tree)  # 假设这是一个在别处定义的函数
        print(self.Blocks,self.Tree)

    def find_circuit(self, leaf, t):
        v = [0]*(self.Branch_Count) # *用于跟踪已被访问过的边
        c = [0]*(self.Junction_Count) # *储存回路中的边，每次发现新的边形成回路时，将它添加到c中
        n = 1   # *记录回路中的边
        s = self.VertexStart[leaf]  # *边起点
        e = self.VertexEnd[leaf]    # *边终点
        c[n] = leaf  # 初始化闭环路节点数组

        sc = False  # 初始化是否找到闭环路的标记
        fnd = False  # 初始化是否找到与当前边相连的边的标记

        while not sc:   # *只要没找到回路就继续循环
            for i in range(self.Junction_Count-1):
                if s == e:
                    sc = True
                    break
                k = t[i]
                if not v[k]:
                    fnd = True
                    if e == self.VertexStart[k]:
                        n += 1
                        c.append(k)
                        e = self.VertexEnd[k]
                    elif e == self.VertexEnd[k]:
                        n += 1
                        c.append(-k)
                        e = self.VertexStart[k]
            if not fnd:  # 如果循环结束时仍未找到与当前边相连的边，则跳出循环
                break
            v[k] = True  # 将已访问的节点标记为True
            fnd = False  # 重置是否找到与当前边相连的边的标记
        return c[:n]  # 返回闭环路节点数组（不包括重复节点）

    def save1(self):
        i = 0
        # shtBranches = openpyxl.load_workbook('/Users/xavier/Documents/Py/找回路等6个文件/实例网络解算.xlsm')
        while i < self.Branch_Count:
            self.shtBranches['Branch'].cell(row=i+1, column=9).value = self.AirQuantities[i]
            i += 1

    def main(self):
        i, j, k, s = 0, 0, 0, 0
        f, f1, dq = 0.0, 0.0, 0.0
        er = 0.001

        self.read()
        self.make_circuits()
        self.initQ()

        while True:
            maxEr = 0.0
            maxEr = max(maxEr, abs(dq))
            for i in range(self.Circuit_Count):
                f = self.circuitF(i)
                f1 = self.circuitF1(i)
                if i == self.Circuit_Count - 1:  # 这里需要根据你的实际情况来设置最后一个电路的处理逻辑
                    f = f - self.fan(self.AirQuantities[self.Branch_Count-1])
                    f1 = self.circuitF1(self.Circuit_Count - 1) - self.fan1(self.AirQuantities[self.Branch_Count-1])
                dq = -f / f1
                if abs(dq) > maxEr:
                    maxEr = abs(dq)
                for j in range(self.BranchCountInCircuit[i]):
                    k = abs(self.BranchNumbersInCircuit[i][j])
                    s = self.BranchNumbersInCircuit[i][j] / abs(self.BranchNumbersInCircuit[i][j])  # 这里需要根据你的实际情况来设置s的计算逻辑
                    self.AirQuantities[k-1] += s * dq
            # 在控制台打印maxEr的值，用于调试        
            print(maxEr)  
            print(self.VertexStart)
            print(self.VertexEnd)
            print(self.AirQuantities)
            print("\n")
            if maxEr < er:  # 如果maxEr小于er，则退出循环
                break
        self.save1()  # 这里需要根据你的实际情况来设置Save1函数的实现


if __name__ == "__main__":
    Mine_Network().main()